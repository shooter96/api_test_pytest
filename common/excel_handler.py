# @author: wy
# @project:api_test_pytest
import time
from openpyxl import load_workbook
import logging
import json
import re


# openpyxl 通用方法封装，读取excel数据，写excel数据
class CaseInfo:
    # 用于存放测试用例
    def __init__(self):
        self.host = None
        self.headers = {}


class ExcelHandler:
    def __init__(self, filename):
        '''
        读取excel,持有整个excel
        '''
        self.filename = filename
        self.work_book = load_workbook(filename)

    def close(self):
        self.close()

    def get_sheet_names(self):
        '''
        拿到当前excel所有的sheet
        返回列表
        '''
        return self.work_book.sheetnames

    def write_excel(self, case_infos, sheet):
        '''
        result: 每执行完一个sheet就会写一个sheet的测试结果
        '''
        # 回写测试结果
        for i in range(len(case_infos)):
            # 获取每一行的用例数据
            case_info = case_infos[i]
            # 获取执行结果
            status = case_info.status
            response_content = case_info.response_content
            response_time = case_info.time_used
            sheet.cell(row=i + 6, column=15, value=str(status))
            sheet.cell(row=i + 6, column=16, value=str(response_content))
            sheet.cell(row=i + 6, column=17, value=str(response_time))
        self.work_book.save(self.filename)
        self.work_book.close()


class CaseInfoHandler:
    def __init__(self, work_book, sheet_name):
        """
        解析当前 sheet 页为 CaseData 对象列表

        work_book: 整个工作薄
        sheet_name：当前sheet页

        """
        # 拿到sheet页每一个格子的数据
        self.sheet = work_book[sheet_name]
        # 获取用例所有的行(获取单元格，cell对象)
        self.rows = list(self.sheet.rows)
        # 获取表头信息
        self.title = [row.value for row in self.rows[4]]
        # 定义用例列表
        self.case_infos = []
        # 获取请求地址信息（获取单元格的值,下标从1开始）
        self.default_host = self.sheet.cell(row=4, column=2).value
        # 构造登录信息
        self.login_info = CaseInfo()
        self.build_login_info()

        # 构造所有用例信息
        self.build_request_info()

    def build_login_info(self):
        self.login_info.params = self.sheet.cell(row=3, column=2).value
        # 登录的地址
        self.login_info.url = self.default_host + self.sheet.cell(row=2, column=2).value
        self.login_info.method = "post"
        self.login_info.headers = {'Content-Type': 'application/json'}
        return self.login_info

    def build_request_info(self):
        """
        将excel的数据组装在一起
        """
        for row in range(5, len(self.rows)):
            # 取整行数据为列表
            case = [column.value for column in self.rows[row]]
            case_info = CaseInfo()
            # 将表头与测试数据打包成元组,返回的是带元组的列表
            for i in zip(self.title, case):
                # 将对应的key，value分别设为对象属性名和属性值(反射机制)
                setattr(case_info, i[0], i[1])
            case_info.host = self.default_host
            case_info.url = None
            # 定义当前用例所在行数(其实该行数为实际对应 Excel 中行数减 1)
            case_info.row = row
            # 将每行生成的用例数据添加到用例列表中
            self.case_infos.append(case_info)

    def parse_params(self, case_info):
        '''
         用于解析当前用例信息中参数需要的占位符

        即将 ex_keys 中的字段，以及 ex_values 中的值解析到 params 中

        '''
        params = getattr(case_info, 'params')
        params = {} if not params else json.loads(params)
        ex_keys = getattr(case_info, 'ex_keys')
        ex_values = getattr(case_info, 'ex_values')
        relay_key = getattr(case_info, 'relay_keys')

        if ex_keys:
            # 先根据 ","分割,可以知道有几个依赖值要处理
            keys = ex_keys.splite(',')
            values = ex_values.splite(',')
            for key in keys:

                # 再根据"&"切割获取依赖的行
                steps = values.splite('&')
                # 获取excel第一列的数据,获取依赖行的关键字
                relay_keys = [column.value for column in self.sheet.columns[1]]
                try:
                    # 获取依赖行的行号
                    row = relay_keys.index(steps[0]) + 1
                    relay_value = self.get_relay_value(row, steps[1])
                    # 将得到的依赖值更新到parmas中
                    self.set_relay_value(params, key, relay_value)
                except ValueError:
                    raise "依赖key{}不存在".format(steps[0])
        setattr(case_info, 'params', params)
        # 返回当前 Handler
        return self

    def parse_path(self, case_info):
        """
        解析url中的占位符
        要求：
        在 ex_keys 中填写对应着 url 占位符的字段
        在 ex_values 中填写取值步骤
        在parse_params已经将所有的依赖值插入params中
        """
        path = getattr(case_info, 'path')
        params = getattr(case_info, 'params')
        # 使用正则获取path中的依赖key
        # compile 函数用于编译正则表达式，生成一个正则表达式（ Pattern ）对象
        pt = re.compile(r'{(.*?)}')
        keys = re.findall(pt, path)  # 得到一个列表[key1,key2]
        for key in keys:
            # value统一转换成字符串
            value = str(params[key])
            if value:
                # 将value填充到path中
                path = path.replace('{' + key + '}', value)
        setattr(case_info, 'path', path)
        # 返回当前 Handler
        return self

    def get_relay_value(self, row, step):
        '''
        row: 依赖的行
        setp: 取值的步骤
        取值
        1、取字典中的值
        2、取列表中的值

        填值：
        1、往字典指定层数添加数据
        2、往字典内的列表中的指定层数添加数据
        '''

        # 拿到依赖的行的内容
        ci = [ci for ci in self.case_infos if getattr(ci, 'row') == int(row)]
        ci = ci[0]
        # 去除response_content格子的内容
        response_content = getattr(ci, 'response_content')
        # 注：excel读取出来的值，数值类型为数值类型，其余的均为str类型
        # 将response_content内容转换为字典类型
        if response_content:
            ci = json.loads(response_content)
        else:
            return None
        from common.utils_handler import get_value
        return get_value(ci, step)

    def set_relay_value(self, params, step, value):
        """
         依赖值所处的位置：
        1、路径参数
        2、body参数
        """
        # 先往body里添加value
        step_list = (step).split('.')
        for key in step_list:
            # 如果为数值类型则人为是列表
            key = key if not key.isdigit() else int(key)
            # 如果为最后一位里则直接插入
            if key == step_list[-1]:
                params[key] = value
            # 插入一个列表,然后在列表里插入字典：
            # driverTableIds.0,driverTableIds.0.driverTableId,driverTableIds.0.driver,driverTableIds.0.config
            # 如果最后一位为数字，则认为是列表
            elif len(step_list) > 1 and step_list[-1].isdigit():
                try:
                    value_list = [] if not params[step_list[-1]] else params[step_list[-1]]
                except KeyError:
                    value_list = []
                value_list.append(value)
                params[step_list[-2]] = value_list
                break
            else:
                try:
                    # 从目标中去key对应的值作为下一次的目标
                    params = params[key]
                except KeyError:
                    # 报错，下一目标为{}
                    params = {}


if __name__ == '__main__':
    pass
    # eh = ExcelHandler('../datas/isc-permission-service/多租户-权限管理.xlsx')
    # work_book = eh.work_book
    #
    # sheet = '平台租户-用户管理'
    # case_infos = CaseInfoHandler(work_book, sheet)
    # # case_info = CaseInfoHandler
    # content = {"code": 0, "message": "成功", "data": [{"driverId": "3853ufdjgyre6y3i"}]}
    # step = "data.0.driverId"
    # value = case_infos.get_value(content, step)
    # print(value)

    # login_info = case_infos.login_info
    # print(list(case_infos.rows))
    # print(case_infos.title)
    # print(case_infos.login_info)
    # print(case_infos.default_host)
    # print(case_infos.login_info.url)
    # params = {"aa": 11, "bb": "hello", "id": 1247823859}
    # path = '/ewyu/wuriwo/{id}'
    # aa = parse_path1(path, params)
    # print(aa)
