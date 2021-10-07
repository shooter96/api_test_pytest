import json
import re
from openpyxl import load_workbook
from openpyxl.styles import Font
from utils.test_utils import str_is_none
from utils.test_utils import get_value
from utils.test_utils import set_value


class CaseInfo:

    def __init__(self):
        """
        构造 CaseInfo 对象，后续所有用例均解析为此对象
        """
        self.expected_key = None
        self.headers = {}


class ExcelParser:

    def __init__(self, filename):
        """
        持有整个 Excel
        """
        self.work_book = load_workbook(filename)

    def get_sheet_names(self):
        """
        拿到当前 Excel 所有的 Sheet
        """
        return self.work_book.sheetnames

    def get_sheet_count(self):
        """
        拿到当前 Excel Sheet 的总数
        """
        return len(self.work_book.sheetnames)


class CaseInfoHolder:

    def __init__(self, work_book, sheet_name, params):
        """
        解析当前 sheet 页为 CaseInfo 对象列表

        :param work_book: 整个工作簿
        :param sheet_name: 当前 sheet 页
        """
        # 取工作簿中对应 sheet 页保存
        self.sheet = work_book[sheet_name]
        # 获取数据总行数
        self.rows = list(self.sheet.rows)
        # 获取表头列表
        self.title = [column.value for column in self.rows[4]]
        # 定义持有的用例列表
        self.case_infos = []
        # 获取登录行信息
        self.default_host = self.sheet.cell(row=4, column=2).value
        if len(params) > 1:
            self.default_host = params[1]
        # 构建登录信息
        self.login_info = CaseInfo()
        self.build_longin_info()
        # 构建所有请求信息
        self.build_request_info()

    def build_longin_info(self):
        """
        构建请求登录的信息
        """
        self.login_info.method = 'post'
        login_host = self.sheet.cell(row=2, column=2).value
        self.login_info.host = login_host
        if login_host and self.default_host:
            temp = login_host.split('/', 3)
            del temp[0]
            del temp[0]
            temp[0] = self.default_host
            self.login_info.host = '/'.join(temp)
        params = self.sheet.cell(row=3, column=2).value
        self.login_info.params = params
        self.login_info.path = ''
        # 请求头为 application/x-www-form-urlencoded
        self.login_info.headers = {'Content-Type': 'application/json'}

    def build_request_info(self):
        """
        构建所有用例信息
        """
        case_count = len(self.rows)
        # 从第六行开始取用例信息
        for row in range(5, case_count):
            # 获取整行信息为列表
            info = [column.value for column in self.rows[row]]
            # 定义当前要处理的用例对象
            case_info = CaseInfo()
            # 将表头字段对应的列作为对象字段，将当前行对应的列信息作为值，填入用例对象中
            for i in zip(self.title, info):
                setattr(case_info, i[0], i[1])
            # 定义当前用例所在行数(其实该行数为实际对应 Excel 中行数减 1)
            case_info.row = row
            # case_info.method = case_info.method.lower()
            # 如果期望值没填则忽略该数据
            expected_code = case_info.expected_key
            # 填充默认 host
            if str_is_none(case_info.host):
                case_info.host = self.default_host
            if expected_code is None or expected_code == '':
                continue
            # 加入用例列表
            self.case_infos.append(case_info)

    def parse_param(self, case_info):
        """
        用于解析当前用例信息中参数需要的占位符

        即将 ex_keys 中的字段，以及 ex_values 中的值解析到 params 中
        """
        # 获取当前用例的 params
        params = getattr(case_info, 'params')
        params = {} if str_is_none(params) else json.loads(params)
        # 取当前用例的 ex_keys
        ex_keys = getattr(case_info, 'ex_keys')
        # 如果有需要处理的占位符
        if ex_keys != '' and ex_keys is not None:
            # 以 , 分割 ex_keys 为列表
            keys = ex_keys.split(",")
            # 以 , 分割 ex_values 为列表
            ex_values = getattr(case_info, 'ex_values').split(",")
            # 循环需要加入的参数字段
            for i in range(0, len(keys)):
                # 不需要拼接入参处理
                if "+" not in ex_values[i]:
                    # 将需要加入的值以 : 分割，得到依赖的行数和取值步骤
                    row_steps = ex_values[i].split(":")
                    # 用对应行和取值步骤去取值
                    value = self.get_value(row_steps[0], row_steps[1])
                    # 取到的值为数字类型的字符串转为 int 类型
                    if isinstance(value, str) and value.isdigit():
                        value = int(value)
                    # 将取出的值插入以对应的字段名插入到 params 中
                    set_value(value, params, keys[i])
                # 需要拼接入参处理
                else:
                    values = ""
                    link_values = ex_values[i].split("+")
                    for j in range(0, len(link_values)):
                        if ":" and "data" in link_values[j]:
                            row_step = link_values[j].split(":")
                            value = self.get_value(row_step[0], row_step[1])
                            values += value
                        else:
                            value = link_values[j]
                            values += value
                    set_value(values, params, keys[i])
        # 处理完成后设置回用例对象中
        setattr(case_info, 'params', params)
        # 返回当前 holder
        return self

    def parse_path(self, case_info):
        """
        用于解析 path 中的占位符

        要求：
        在 ex_keys 中填写对应着 url 占位符的字段
        在 ex_values 中填写取值步骤
        """
        # 取当前用例对象的 path 字段
        path = getattr(case_info, 'path')
        # 取当前用例对象中的 params 字段
        params = getattr(case_info, 'params')
        # 正则取占位符对应的字段
        pt = re.compile(r'[{](.*?)[}]')
        keys = re.findall(pt, path)
        # 如果没有占位符直接返回
        if len(keys) == 0:
            return
        # 存在占位符则循环处理
        for key in keys:
            # 从 params 中取出占位符字段对应的值
            value = params[key]
            # 不是字符串的话将其转为字符串
            if not isinstance(value, str):
                value = str(value)
            # 将 path 中的 {key} 替换为 value
            if value:
                path = path.replace('{' + key + '}', value)
        # 处理完成后设置回用例对象中
        setattr(case_info, 'path', path)
        # 返回当前 holder
        return self

    def get_value(self, row, steps):
        """
        从给定行数和取值步骤，取出需要的值
        """
        # 取依赖的行数据
        ci = [ci for ci in self.case_infos if getattr(ci, 'row') == int(row)]
        ci = ci[0]
        # 没有取到返回 None
        if not ci:
            return None
        # 否则取出依赖行数据中的请求响应结果
        content = getattr(ci, 'response_content')
        # 有值转为字典
        if content:
            ci = json.loads(content)
        # 无值直接返回 None
        else:
            return None
        # 以给定对象和取值步骤去取出对应数据
        return get_value(ci, steps)


class ExcelWriter:
    def __init__(self, filename, work_book, sheet):
        """
        初始化需要写入的 Excel
        """
        self.work_book = work_book
        self.filename = filename
        self.sheet = sheet

    def write(self, case_infos, log_info, default_host):
        """
        开始写结果
        """
        # 写结果信息
        self.set_columns(case_infos, log_info, default_host)
        # 保存结果
        self.work_book.save(self.filename)
        # 关闭 Excel
        self.work_book.close()

    def set_columns(self, case_infos, log_info, default_host):
        """
        写结果信息
        """
        # 遍历写所有用例信息
        for i in range(0, len(case_infos)):
            # 取用例对象信息
            case_info = case_infos[i]
            # 取用例请求结果
            status = case_info.status
            # 失败：结果 Excel 为红色
            if status == 'failed':
                self.sheet.cell(row=6 + i, column=15, value=status).font = Font(bold=True, color='FF0000')
            # 未执行，默认颜色
            elif status == '未执行':
                self.sheet.cell(row=6 + i, column=15, value=status).font = Font(bold=True, color='000000')
            # 成功：结果 Excel 为绿色
            else:
                self.sheet.cell(row=6 + i, column=15, value=status).font = Font(bold=True, color='3CB371')
            # 写请求的实际 code
            self.sheet.cell(row=6 + i, column=16, value=case_info.response_code)
            # 写请求的返回值
            self.sheet.cell(row=6 + i, column=17, value=str(case_info.response_content))
            # 写请求耗时
            self.sheet.cell(row=6 + i, column=18, value=str(case_info.time_used))
        self.sheet.cell(row=2, column=2, value=log_info.host)
        self.sheet.cell(row=4, column=2, value=default_host)
